import os
import logging
import threading
import time
from slack_sdk import WebClient
from slack_sdk.socket_mode import SocketModeClient
from slack_sdk.socket_mode.response import SocketModeResponse
from slack_sdk.errors import SlackApiError
from dotenv import load_dotenv
from datetime import datetime

# For excel writing
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from utils.mail_function import schedule_email

# Configure logging
logging.basicConfig(level=logging.INFO)  # Set to DEBUG for more details
logger = logging.getLogger(__name__)

# Load environment variables from .env file
load_dotenv('../.env')

# Initialize Slack clients
slack_bot_token = os.getenv("SLACK_BOT_TOKEN")
slack_app_token = os.getenv("SLACK_APP_TOKEN")

web_client = WebClient(token=slack_bot_token)
socket_client = SocketModeClient(app_token=slack_app_token, web_client=web_client)

def get_user_name(user_id):
    """
    Retrieve the user's display name from Slack using their user ID.
    """
    try:
        response = web_client.users_info(user=user_id)
        if response["ok"]:
            return response["user"]["profile"]["real_name"] or response["user"]["name"]
        else:
            print(f"Failed to fetch user info: {response['error']}")
    except SlackApiError as e:
        print(f"Error fetching user info: {e.response['error']}")
    return "Unknown User"

# Function to initialize the Excel file if it doesn't exist
def initialize_excel_file(path):
    # Format the filename as "DD-MM_CIT.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Check in CAG"
    # Write headers
    sheet.append(["User", "Timestamp"])
    workbook.save(path)


def check_duplicate_user(excel_file, user):
    # Load the workbook and access the active sheet
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    # Check if the user already exists in the "User" column
    user_column_index = 1  # Assuming "User" is in the first column
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=user_column_index, max_col=user_column_index):
        for cell in row:
            if cell.value == user:
                logger.info(f"Duplicate user found: {user}")
                return True

    logger.info(f"No duplicate found for user: {user}")
    return False



def handle_message(event_data):
    message = event_data.get("text", "")
    user = event_data.get("user", "")
    channel = event_data.get("channel", "")
    user_name = get_user_name(user)

    if "checkin" in message.lower() or "check in" in message.lower():
        # Write to an excel instead
        row_data = [user_name, datetime.utcnow().strftime("%H:%M")]
        # Send confirmation to the channel
        # This will determine file path
        current_date = datetime.now().strftime("%d-%m_CIT")
        excel_file = f"checkin_records/{current_date}.xlsx"
        if not os.path.exists(excel_file):
            initialize_excel_file(excel_file)
        elif check_duplicate_user(excel_file, user_name):
            # return error message
            web_client.chat_postMessage(
                channel=channel,
                text=f"Hi @{user_name}. You have already checked in for today! Please do not send duplicate messages 😡."
            )
            return
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        sheet.append(row_data)  # Append the row
        workbook.save(excel_file)  # Save changes
        logger.info(f"Check in for @{user_name} has been recorded in the excel")
        try:
            web_client.chat_postMessage(
                channel=channel,
                text=f"@{user_name} has checked in! Have a wonderful day at the office today!."
            )
            logger.info(f"Check in for @{user_name} is successful")
        except SlackApiError as e:
            logger.info(f"Error posting message: {e.response['error']}")
    else:
        web_client.chat_postMessage(
            channel=channel,
            text=f"Hi @{user_name}. I do not understand that message. Please send only either 'Check In' or 'Checkin'."
        )
        logger.info(f"@{user_name} sent an invalid message")
# Function to handle app mentions for "show notes"
# def handle_app_mention(event_data):
#     text = event_data.get("text", "")
#     channel = event_data.get("channel", "")
#
#     if "" in text.lower():
#         notes_summary = "\n".join(
#             [f"{i+1}. <@{entry['user']}>: \"{entry['text']}\"" for i, entry in enumerate(message_log)]
#         ) if message_log else "No notes recorded yet."
#
#         try:
#             web_client.chat_postMessage(
#                 channel=channel,
#                 text=f"Here are the recorded notes:\n{notes_summary}"
#             )
#         except SlackApiError as e:
#             logger.info(f"Error posting message: {e.response['error']}")


# Handle events from the Socket Mode client
def handle_events(client, payload):
    # Log the payload for debugging
    # logger.debug(f"Received payload: {payload}")

    if payload.type == "events_api":
        event = payload.payload.get("event", {})
        logger.info(f"Processing event: {event}")

        if event.get("type") == "message" and not event.get("bot_id"):  # Ignore bot messages
            handle_message(event)
        # elif event.get("type") == "app_mention":
        #     handle_app_mention(event)

        # Acknowledge the event
        client.send_socket_mode_response(SocketModeResponse(envelope_id=payload.envelope_id))


# Start the Socket Mode client
if __name__ == "__main__":
    socket_client.socket_mode_request_listeners.append(handle_events)
    logger.info("⚡️ Slack Bot is running!")

    # Connect to Slack
    socket_client.connect()
    # Schedule the email to be sent at 23:59 Singapore Time (SGT)

    # Run the scheduler in a separate thread
    scheduler_thread = threading.Thread(target=schedule_email, daemon=True)
    scheduler_thread.start()

    # Prevent the script from exiting
    try:
        while True:
            time.sleep(1)  # Keep the script alive
    except KeyboardInterrupt:
        logger.info("Bot stopped manually.")